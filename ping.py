import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import openpyxl
import subprocess
import socket

# Discord-Farben als Konstanten
DISCORD_BG = "#2C2F33"   # Grund-Hintergrund
DISCORD_DARKER_BG = "#23272A"
DISCORD_ACCENT = "#5865F2"  # Blurple
DISCORD_ACCENT_HOVER = "#4752C4"
DISCORD_TEXT = "#FFFFFF"
DISCORD_TEXT_MUTED = "#99AAB5"

def main():
    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("Discord-Style IP Checker")
    root.geometry("700x500")  # Startgröße des Fensters
    root.minsize(600, 400)    # Minimale Fenstergröße
    root.configure(bg=DISCORD_BG)

    # -- Style für ttk-Widgets erstellen --
    style = ttk.Style(root)

    # Neuen Theme-Namen erstellen (z. B. "Discord")
    style.theme_create("Discord", parent="alt", settings={
        "TFrame": {
            "configure": {
                "background": DISCORD_BG
            }
        },
        "TLabel": {
            "configure": {
                "foreground": DISCORD_TEXT,
                "background": DISCORD_BG,
                "font": ("Helvetica", 10)
            }
        },
        "TButton": {
            "configure": {
                "foreground": DISCORD_TEXT,
                "background": DISCORD_ACCENT,
                "font": ("Helvetica", 10),
                "borderwidth": 0,
                "padding": 6
            },
            "map": {
                "background": [
                    ("active", DISCORD_ACCENT_HOVER),
                    ("disabled", DISCORD_TEXT_MUTED)
                ],
                "foreground": [
                    ("disabled", DISCORD_BG)
                ]
            }
        },
        "Horizontal.TProgressbar": {
            "configure": {
                "troughcolor": DISCORD_DARKER_BG,
                "background": DISCORD_ACCENT,
                "bordercolor": DISCORD_BG,
                "lightcolor": DISCORD_ACCENT,
                "darkcolor": DISCORD_ACCENT,
                "thickness": 10
            }
        }
    })

    style.theme_use("Discord")

    # -- Frames für besseres Layout --
    # Oberer Bereich (Label und Button)
    top_frame = ttk.Frame(root)
    top_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=(10, 5))

    # Fortschrittsbalken
    progress_frame = ttk.Frame(root)
    progress_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)

    # Text-Ausgabefeld
    bottom_frame = ttk.Frame(root)
    bottom_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(5, 10))

    # -- Responsive Konfiguration --
    root.rowconfigure(2, weight=1)  # Das Textfeld (unten) soll sich ausdehnen
    root.columnconfigure(0, weight=1)
    top_frame.columnconfigure(0, weight=1)
    progress_frame.columnconfigure(0, weight=1)
    bottom_frame.columnconfigure(0, weight=1)
    bottom_frame.rowconfigure(0, weight=1)

    # -- Widgets im top_frame --
    label = ttk.Label(top_frame, text="Wähle eine Excel-Datei (Spalte A=IP, ab Zeile 2).")
    label.grid(row=0, column=0, sticky="w", padx=5, pady=5)

    button_select_file = ttk.Button(top_frame, text="Excel-Datei auswählen")
    button_select_file.grid(row=0, column=1, sticky="e", padx=5, pady=5)

    # -- Widgets im progress_frame --
    progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", mode='determinate')
    progress_bar.grid(row=0, column=0, sticky="ew")

    # -- Widgets im bottom_frame --
    # Für das Textwidget (tk.Text) müssen wir manuell Farben setzen,
    # da es kein ttk-Widget ist.
    text_result = tk.Text(bottom_frame, bg=DISCORD_DARKER_BG, fg=DISCORD_TEXT, wrap="word")
    text_result.grid(row=0, column=0, sticky="nsew")

    # Scrollbar
    scrollbar = ttk.Scrollbar(bottom_frame, orient="vertical", command=text_result.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    text_result.configure(yscrollcommand=scrollbar.set)

    # -- Funktion zum Auswählen der Datei und Bearbeiten --
    def select_file():
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel-Dateien", "*.xlsx *.xls"), ("Alle Dateien", "*.*")]
        )
        if not file_path:
            return

        # Textfeld leeren
        text_result.delete("1.0", tk.END)

        # Excel öffnen
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # Erstes Tabellenblatt

        max_row = ws.max_row
        total_ips = max_row - 1 if max_row > 1 else 0

        # Fortschrittsbalken zurücksetzen
        progress_bar["maximum"] = total_ips
        progress_bar["value"] = 0

        for row_index in range(2, max_row + 1):
            ip_val = ws.cell(row=row_index, column=1).value
            if not ip_val:
                continue  # Nächste Zeile, falls keine IP

            ip_str = str(ip_val).strip()
            if not ip_str:
                continue

            # -- Ping --
            ping_result = subprocess.run(
                ["ping", "-n", "1", "-w", "1000", ip_str],
                stdout=subprocess.DEVNULL
            )
            if ping_result.returncode == 0:
                ws.cell(row=row_index, column=2).value = "ja"
                ping_status = "ja"
            else:
                ws.cell(row=row_index, column=2).value = "nein"
                ping_status = "nein"

            # -- DNS-Lookup --
            try:
                hostname, _, _ = socket.gethostbyaddr(ip_str)
                ws.cell(row=row_index, column=3).value = hostname
                dns_status = hostname
            except Exception:
                ws.cell(row=row_index, column=3).value = "nein"
                dns_status = "nein"

            # -- Ausgabe ins Textfeld --
            text_result.insert(tk.END, f"{ip_str}: Ping={ping_status}, DNS={dns_status}\n")

            # -- Fortschritt aktualisieren --
            progress_bar["value"] += 1
            progress_bar.update()

        # Änderungen speichern
        wb.save(file_path)
        text_result.insert(tk.END, "\nErgebnisse wurden in die Excel-Datei geschrieben.\n")

    # Button-Funktion verbinden
    button_select_file.config(command=select_file)

    # Eventloop starten
    root.mainloop()


if __name__ == "__main__":
    main()
